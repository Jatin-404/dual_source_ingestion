from fastapi import APIRouter, BackgroundTasks, UploadFile, File, Form, HTTPException
from pydantic import BaseModel
from httpx import AsyncClient
import uuid
import asyncio
import json
import zipfile
import io
import os

# PDF and DOCX readers
import fitz                          # pymupdf -> for PDFs
from docx import Document as DocxDoc # python-docx -> for DOCX


router = APIRouter(prefix='/ingest', tags=["Ingestion"])

CHUNKER_URL = "http://localhost:8001"
EMBED_URL   = "http://localhost:8002"
STORE_URL   = "http://localhost:8003"

jobs = {}

# MODELS

class CaseMetadata(BaseModel):
    case_number: str                
    date: str                       
    court: str                      
    lawyers: list[str] 

class IngestRequest(BaseModel):
    content_text: str               
    metadata: CaseMetadata          # case info → stamped on every chunk
    document_id: str | None = None
    filename: str                   
    chunk_size: int = 500

# FILE TEXT EXTRACTION

def extract_text_from_file(filename: str, file_bytes: bytes) -> str:
    """
    Reads txt, pdf, or docx and returns plain text.
    This is why we installed pymupdf and python-docx.
    """
    ext = os.path.splitext(filename)[1].lower()

    if ext == ".txt":
        return file_bytes.decode("utf-8")

    elif ext == ".pdf":
        # fitz opens from bytes in memory - no temp file needed
        pdf = fitz.open(stream=file_bytes, filetype = "pdf")
        text = ''
        for page in pdf:
            text += page.get_text()

        return text.strip()
    
    elif ext == ".docx":
        # python-docx needs a file-like obejct
        doc = DocxDoc(io.BytesIO(file_bytes))
        text = "\n".join([para.text for para in doc.paragraphs])
        return text.strip()
    
    else:
        raise ValueError(f"Unsupported file type: {ext}. Supported: .txt, .pdf, .docx")
    
def parse_metadata(metadata_bytes: bytes, filename: str) -> CaseMetadata:
    """
    Reads metadata from either JSON or XLSX.
    """
    ext = os.path.splitext(filename)[1].lower()

    if ext == ".json":
        try:
            data = json.loads(metadata_bytes.decode("utf-8"))
            # lawyers might come as comma-separated string from Excel
            if isinstance(data.get("lawyers"), str):
                data["lawyers"] = [l.strip() for l in data["lawyers"].split(",")]
            return CaseMetadata(**data)
        except Exception as e:
            raise ValueError(f"Invalid metadata JSON: {str(e)}")
        
    elif ext == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(metadata_bytes))
        ws = wb.active
        # First row = headers, secod row = values
        headers = [cell.value for cell in ws[1]]
        values = [cell.value for cell in ws[2]]
        data = dict(zip(headers,values))

        # lawyers column: "Advocate A, Advocate B" -> list

        if isinstance(data.get("lawyers"), str):
            data["lawyers"] = [l.strip() for l in data["lawyers"].split(",")]

        return CaseMetadata(**data)
    
    else:
        raise ValueError(f"Unsupported metadata format: {ext}. Use .json or .xlsx")




# Core Pipeline


async def ingest_one(request: IngestRequest, client: AsyncClient):

    chunk_response = await client.post(f"{CHUNKER_URL}/chunk",
                                        json={
                                            "content_text": request.content_text,
                                            "metadata": request.metadata.model_dump(),
                                            "filename": request.filename,
                                            "chunk_size": request.chunk_size
                                        }) 
    chunk_data = chunk_response.json()
    chunks = chunk_data["chunks"]
    document_id = chunk_data["document_id"]


    embed_response = await client.post(f"{EMBED_URL}/embed",
                                        json={
                                            "chunks" : chunks
                                        })
    
    embed_chunks = embed_response.json()["chunks"]


    await client.post(f"{STORE_URL}/store",
                        json={"chunks": embed_chunks})

    return {
        "document_id": document_id,
        "total_chunks": len(chunks),
        "case_number": request.metadata.case_number,
        "court": request.metadata.court
    }
    
async def run_ingest_batch(requests: list[IngestRequest], job_id: str):
    jobs[job_id] = {"status": "processing"}
    try:
        async with AsyncClient(timeout= 60) as client:
            jobs[job_id] = {"status": "processing"}
            # run all cases in parallel
            tasks = [ingest_one(req, client) for req in requests]
            results = await asyncio.gather(*tasks, return_exceptions=True)

        # separte success from failures

        completed = []
        failed = []
        for req , result in zip(requests,results):
            if isinstance(result, Exception):
                failed.append({
                    "case_number": req.metadata.case_number,
                    "error": str(result)
                })
            else:
                completed.append(result)
        jobs[job_id] = {
            "status": "completed",
            "total_cases": len(requests),
            "succeeded": len(completed),
            "failed": len(failed),
            "results": completed,
            "errors": failed
        }

    except Exception as e:
        jobs[job_id] = {"status": "failed", "error": str(e)}


def parse_zip(zip_bytes: bytes, chunk_size: int) -> tuple[list[IngestRequest], list[dict]]:
    """
    Opens the zip, iterates folders, finds metadata + content file in each.
    Returns (valid_requests, errors)

    Expected structure:
    upload.zip
    ├── case_001/
    │     ├── metadata.json   (or metadata.xlsx)
    │     └── file.txt        (or file.pdf / file.docx)
    """
    valid_requests = []
    errors = []

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        all_paths = zf.namelist()

        # Group files by their IMMEDIATE parent folder
        # Handles both:
        #   case_001/file.txt          (zipped correctly)
        #   upload/case_001/file.txt   (zipped from parent folder)
        folders: dict[str, list[str]] = {}
        for path in all_paths:
            parts = [p for p in path.split("/") if p]  # remove empty strings
            
            if len(parts) < 2:
                continue   # skip root-level loose files

            # file is always last, its parent folder is second to last
            parent_folder = parts[-2]   # ← always the immediate parent
            full_folder_path = "/".join(parts[:-1])  # for reading later

            if full_folder_path not in folders:
                folders[full_folder_path] = []
            folders[full_folder_path].append(path)

        for folder, file_paths in folders.items():
            # skip if this folder only contains subfolders (no actual files)
            actual_files = [p for p in file_paths 
                          if not p.endswith("/")]
            if not actual_files:
                continue

            try:
                metadata_path = None
                content_path  = None

                for path in actual_files:
                    fname = os.path.basename(path).lower()
                    ext   = os.path.splitext(fname)[1]

                    if fname.startswith("metadata") and ext in (".json", ".xlsx"):
                        metadata_path = path
                    elif ext in (".txt", ".pdf", ".docx"):
                        content_path = path

                if not metadata_path:
                    raise ValueError("No metadata.json or metadata.xlsx found")
                if not content_path:
                    raise ValueError("No content file (.txt/.pdf/.docx) found")

                metadata_bytes = zf.read(metadata_path)
                content_bytes  = zf.read(content_path)

                metadata     = parse_metadata(metadata_bytes, os.path.basename(metadata_path))
                content_text = extract_text_from_file(os.path.basename(content_path), content_bytes)

                valid_requests.append(IngestRequest(
                    content_text=content_text,
                    metadata=metadata,
                    filename=os.path.basename(content_path),
                    chunk_size=chunk_size
                ))

            except Exception as e:
                errors.append({"folder": folder, "error": str(e)})

    return valid_requests, errors




# ENDPOINTS

@router.post('/')
async def ingest(requests: list[IngestRequest], background_tasks: BackgroundTasks):
    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "queued"}

    background_tasks.add_task(run_ingest_batch, requests, job_id)
    return{
        "job_id": job_id,
        "status": "queued",
        "msg": "ingestion started in background"
    }

@router.post("/upload")
async def ingest_upload(
    background_tasks: BackgroundTasks,
    content_file: UploadFile = File(...),
    metadata_file: UploadFile = File(...),
    chunk_size: int = Form(default=500)  # optional sent as form field
    ):
    # step-1 reading both files 
    content_bytes = await content_file.read()           # without await  we would get a coroutine object here 
    metadata_bytes = await metadata_file.read()         # the info stored after reading is in bytes

    # step-2 decode content file as plain text

    try: 
        content_text = extract_text_from_file(content_file.filename, content_bytes)
    except ValueError as e:
        raise HTTPException(status_code=400, detail= str(e))
    
    # step-3 parsing metadata file as json

    try:
        metadata = parse_metadata(metadata_bytes, metadata_file.filename)
    
    except ValueError as e:
        raise HTTPException(status_code=400, detail= str(e))
       
    
    # step-4 Build ingest request (same shape as before)

    request = IngestRequest(
        content_text=content_text,
        metadata=metadata,
        filename=content_file.filename,
        chunk_size=chunk_size
    )

    # step 5 handoff to same background pipeline

    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "queued"}
    background_tasks.add_task(run_ingest_batch, [request], job_id)

    return{
        "job_id": job_id,
        "status": "queued",
        "filename": content_file.filename,
        "metadata_file": metadata_file.filename,
        "msg": "File ingestion started in background"  
    }

# ZIP BATCH UPLOAD (multiple cases at once)
@router.post("/batch")
async def ingest_batch_zip(
        background_tasks: BackgroundTasks,
        zip_file: UploadFile = File(...),
        chunk_size: int = Form(default=500)
    ):
    # validate its actually a zip
    if not zip_file.filename.endswith(".zip"):
        raise HTTPException(status_code=400, detail="File must be a .zip")

    zip_bytes = await zip_file.read()

    # Parse zip -> list of IngestRequests
    try:
        valid_requests, parse_errors = parse_zip(zip_bytes, chunk_size)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read zip: {str(e)}")
    
    if not valid_requests:
        raise HTTPException(
            status_code=400,
            detail=f"No valid cases found in zip. Errors: {parse_errors}"
        )
    
    job_id = str(uuid.uuid4())
    jobs[job_id] = {"status": "queued"}
    background_tasks.add_task(run_ingest_batch, valid_requests,job_id)

    return{
        "job_id": job_id,
        "status": "queued",
        "total_cases_found": len(valid_requests),
        "skipped_folders": parse_errors,   # folders that had missing/bad files
        "msg": "ZIP batch ingestion started in background"
    }


@router.get('/jobs/{job_id}')
async def get_job(job_id : str):
    if job_id not in jobs:
        return{"error": "job not found"}
    else:
        return{
            "job_id": job_id,
            **jobs[job_id]
        }
    